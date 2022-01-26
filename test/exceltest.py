"""
操作excel表格
"""
import openpyxl
from openpyxl import worksheet

file_xls_name = "cases.xlsx"
sheet_name = 'test_case'

# 创建excel文件
# wb = openpyxl.Workbook()
# wb.create_sheet(sheet_name)
# wb.save(file_xls_name)

wb = openpyxl.load_workbook(file_xls_name)
sh = wb[sheet_name]
print("sh：", sh)

# rows_data = list(sh.rows())
# columns_data = list(sh.columns)
# print(rows_data)
# print(columns_data)

ce = sh.cell(row=1, column=1)
print("cell：", ce)
print("读取第一行第一列：", ce.value)

wb.close()


